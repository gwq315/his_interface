"""
数据导入导出API路由模块

本模块提供数据导入导出的功能：
1. 导出JSON格式：GET /api/import-export/export/json
2. 导出Excel格式：GET /api/import-export/export/excel
3. 导入JSON格式：POST /api/import-export/import/json（待实现）
4. 导入Excel格式：POST /api/import-export/import/excel（待实现）

导出功能会将所有接口、参数、字典数据导出为文件供下载。
导入功能可用于批量导入数据（待完善）。

作者: Auto
创建时间: 2024
"""

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from backend.database import get_db
from backend.app.crud import get_interfaces, get_dictionaries

# 创建API路由器，所有路由的前缀为/api/import-export
router = APIRouter(prefix="/api/import-export", tags=["导入导出"])


@router.get("/export/json")
def export_json(db: Session = Depends(get_db)):
    """
    导出JSON格式数据
    
    导出所有接口、参数和字典数据为JSON格式文件。
    文件包含完整的接口信息，包括所有参数和关联的字典。
    
    Args:
        db: 数据库会话（自动注入）
        
    Returns:
        FileResponse: JSON文件下载响应
        
    文件格式:
        {
            "interfaces": [...],  # 接口列表
            "dictionaries": [...],  # 字典列表
            "export_time": "2024-01-01T00:00:00"  # 导出时间
        }
    """
    # 获取所有接口和字典数据（最多10000条，通常足够）
    interfaces = get_interfaces(db, skip=0, limit=10000)
    dictionaries = get_dictionaries(db, skip=0, limit=10000)
    
    # 构建导出数据字典
    data = {
        # 接口列表，包含所有接口信息和参数
        "interfaces": [
            {
                "code": iface.code,
                "name": iface.name,
                "description": iface.description,
                "interface_type": iface.interface_type.value,  # 枚举值转为字符串
                "url": iface.url,
                "method": iface.method,
                "category": iface.category,
                "tags": iface.tags,
                "status": iface.status,
                # 接口的所有参数（入参和出参）
                "parameters": [
                    {
                        "name": param.name,
                        "field_name": param.field_name,
                        "data_type": param.data_type,
                        "param_type": param.param_type.value,  # input或output
                        "required": param.required,
                        "default_value": param.default_value,
                        "description": param.description,
                        "example": param.example,
                        "order_index": param.order_index
                    }
                    for param in iface.parameters
                ]
            }
            for iface in interfaces
        ],
        # 字典列表，包含所有字典和字典值
        "dictionaries": [
            {
                "code": dict.code,
                "name": dict.name,
                "description": dict.description,
                # 字典的所有值（键值对）
                "values": [
                    {
                        "key": value.key,
                        "value": value.value,
                        "description": value.description,
                        "order_index": value.order_index
                    }
                    for value in dict.values
                ]
            }
            for dict in dictionaries
        ],
        # 导出时间戳
        "export_time": datetime.now().isoformat()
    }
    
    # 生成文件名（包含时间戳，避免覆盖）
    filename = f"his_interfaces_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    
    # 确保data目录存在（用于存储导出文件）
    data_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "data")
    os.makedirs(data_dir, exist_ok=True)
    filepath = os.path.join(data_dir, filename)
    
    # 写入JSON文件（使用UTF-8编码，确保中文正确显示）
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)  # indent=2使JSON格式更易读
    
    # 返回文件下载响应
    return FileResponse(
        filepath,
        filename=filename,
        media_type="application/json"
    )


@router.get("/export/excel")
def export_excel(db: Session = Depends(get_db)):
    """
    导出Excel格式数据
    
    将所有接口、参数和字典数据导出为Excel格式文件。
    Excel文件包含多个工作表：
    1. 接口列表：所有接口的基本信息
    2. 参数列表：所有接口的参数信息
    3. 字典列表：所有字典的基本信息
    4. 字典值：所有字典的键值对
    
    Args:
        db: 数据库会话（自动注入）
        
    Returns:
        FileResponse: Excel文件下载响应
    """
    # 获取所有接口和字典数据
    interfaces = get_interfaces(db, skip=0, limit=10000)
    dictionaries = get_dictionaries(db, skip=0, limit=10000)
    
    # 创建新的Excel工作簿
    wb = Workbook()
    
    # ========== 工作表1：接口列表 ==========
    ws_interfaces = wb.active
    ws_interfaces.title = "接口列表"
    
    # 定义表头样式（蓝色背景，白色粗体文字）
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # 设置接口列表表头
    headers = ["接口编码", "接口名称", "接口类型", "URL", "方法", "分类", "状态", "描述"]
    ws_interfaces.append(headers)
    # 应用表头样式
    for cell in ws_interfaces[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入接口数据
    for iface in interfaces:
        ws_interfaces.append([
            iface.code,
            iface.name,
            "API接口" if iface.interface_type.value == "api" else "视图接口",  # 转换为中文
            iface.url or "",
            iface.method or "",
            iface.category or "",
            iface.status,
            iface.description or ""
        ])
    
    # ========== 工作表2：参数列表 ==========
    ws_params = wb.create_sheet("参数列表")
    headers = ["接口编码", "参数类型", "字段名", "参数名称", "数据类型", "必填", "默认值", "描述", "示例"]
    ws_params.append(headers)
    # 应用表头样式
    for cell in ws_params[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入参数数据
    for iface in interfaces:
        for param in iface.parameters:
            ws_params.append([
                iface.code,  # 关联的接口编码
                "入参" if param.param_type.value == "input" else "出参",  # 转换为中文
                param.field_name,
                param.name,
                param.data_type,
                "是" if param.required else "否",  # 布尔值转换为中文
                param.default_value or "",
                param.description or "",
                param.example or ""
            ])
    
    # ========== 工作表3：字典列表 ==========
    ws_dicts = wb.create_sheet("字典列表")
    headers = ["字典编码", "字典名称", "描述"]
    ws_dicts.append(headers)
    # 应用表头样式
    for cell in ws_dicts[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入字典数据
    for dict_item in dictionaries:
        ws_dicts.append([
            dict_item.code,
            dict_item.name,
            dict_item.description or ""
        ])
    
    # ========== 工作表4：字典值 ==========
    ws_dict_values = wb.create_sheet("字典值")
    headers = ["字典编码", "键", "值", "描述"]
    ws_dict_values.append(headers)
    # 应用表头样式
    for cell in ws_dict_values[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入字典值数据
    for dict_item in dictionaries:
        for value in dict_item.values:
            ws_dict_values.append([
                dict_item.code,  # 关联的字典编码
                value.key,
                value.value,
                value.description or ""
            ])
    
    # ========== 自动调整列宽 ==========
    # 遍历所有工作表，根据内容自动调整列宽
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            # 找到该列中最长的内容
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # 设置列宽（最小宽度+2，最大50）
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # ========== 保存文件 ==========
    # 生成文件名（包含时间戳）
    filename = f"his_interfaces_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    # 确保data目录存在
    data_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "data")
    os.makedirs(data_dir, exist_ok=True)
    filepath = os.path.join(data_dir, filename)
    
    # 保存Excel文件
    wb.save(filepath)
    
    # 返回文件下载响应
    return FileResponse(
        filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.post("/import/json")
def import_json(db: Session = Depends(get_db)):
    """
    导入JSON格式数据（待实现）
    
    此功能将允许用户上传JSON文件，批量导入接口、参数和字典数据。
    导入时应检查数据完整性，避免重复导入。
    
    Args:
        db: 数据库会话（自动注入）
        
    Raises:
        HTTPException 501: 功能未实现
        
    计划功能:
        - 解析JSON文件
        - 验证数据格式
        - 检查重复数据
        - 批量插入数据库
        - 返回导入结果统计
    """
    # TODO: 实现JSON导入功能
    raise HTTPException(status_code=501, detail="JSON导入功能待实现")


@router.post("/import/excel")
def import_excel(db: Session = Depends(get_db)):
    """
    导入Excel格式数据（待实现）
    
    此功能将允许用户上传Excel文件，批量导入接口、参数和字典数据。
    需要解析多个工作表，并按顺序导入数据。
    
    Args:
        db: 数据库会话（自动注入）
        
    Raises:
        HTTPException 501: 功能未实现
        
    计划功能:
        - 解析Excel文件（支持多个工作表）
        - 验证数据格式
        - 检查重复数据
        - 批量插入数据库
        - 返回导入结果统计
    """
    # TODO: 实现Excel导入功能
    raise HTTPException(status_code=501, detail="Excel导入功能待实现")
